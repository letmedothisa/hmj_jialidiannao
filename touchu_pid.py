# -*- coding:utf-8 -*-
import datetime
import tkinter
import tkinter.messagebox
import os
import openpyxl
import mysql_tengxun
import time
import serial

data_zaoyan = []
data_zaoyan_cunchu = []
li_fix=[]
li_dcf=[]
from tkinter import ttk
from openpyxl import load_workbook
li_diancifa = []
# 获取拍照的存储时间以及图片的存储路径
with open('config.txt', 'r') as f:
    a = f.readlines()
    for i in range(len(a)):
        b = a[i].split('_____')
        if b[0] == 'control_dcf':
            li_diancifa_1 = []
            li_diancifa_1.append(b[1].replace("\n", ""))
            li_diancifa_1.append(b[2].replace("\n", ""))
            li_diancifa_1.append(b[3].replace("\n", ""))
            li_diancifa_1.append(b[4].replace("\n", ""))
            li_diancifa_1.append(b[5].replace("\n", ""))
            li_diancifa_1.append(b[6].replace("\n", ""))
            li_diancifa_1.append(b[7].replace("\n", ""))
            li_diancifa.append(li_diancifa_1)



def creat_excel(excel_filename):
    if os.path.exists(excel_filename):
        os.remove(excel_filename)
        time.sleep(0.5)

    # 生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
    wb = openpyxl.Workbook()
    # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
    ws = wb.active
    # 更改工作表ws的title
    ws.title = '点火控制记录'
    ws['A1'] = '灶眼'
    ws['B1'] = '点火时间'
    ws['C1'] = '点火时长'
    ws['A2'] = '0'
    ws['A3'] = '1'
    ws['A4'] = '2'
    ws['A5'] = '3'
    ws['A6'] = '4'
    ws['A7'] = '5'
    ws['A8'] = '6'
    ws['A9'] = '7'
    ws['A10'] = '8'
    ws['A11'] = '9'
    ws['A12'] = '10'
    ws['A13'] = '11'
    wb.save(excel_filename)

def operate(operate_dcf):
    ser = serial.Serial("COM3", 9600, 8)
    len_return_data = 0
    if ser.isOpen():
        result = ser.write(bytes.fromhex(operate_dcf))
        print(result)
        time.sleep(0.5)
        len_return_data = ser.inWaiting()

    ser.close()
    return len_return_data

def contorl_dcf(excel_filename,li_diancifa):
    print('control-diancifa')
    if len(li_dcf)>0:
        number=li_dcf[0]
        excel = load_workbook(excel_filename)
        # 获取sheet：
        table = excel['点火控制记录']  # 通过表名获取
        if table['B' + str(number + 2)].value:
            count = operate(li_diancifa[number][2])
            if count != 0:
                yaji_long = round((datetime.datetime.now() - table['B' + str(number + 2)].value).seconds / 60, 2)

                sql = "update shangqi_2 set yaji_long = '%s',guanhuo_time ='%s' where yaji_time = '%s'" % (
                    yaji_long, datetime.datetime.now(), table['B' + str(number + 2)].value)
                try:
                    mysql_tengxun.mysql_db(sql)
                except:
                    time.sleep(0.5)
                    with open('diancifa_error.txt', 'a') as f:
                        f.write(
                            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '_update_data_error_' + '\n')

        elif not table['B' + str(number + 2)].value:
            count = operate(li_diancifa[number][1])
            if count != 0:
                dianhuo_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                dianhuo_time = datetime.datetime.strptime(dianhuo_time, '%Y-%m-%d %H:%M:%S')
                table['B' + str(number + 2)] = dianhuo_time
                print('ok')
                sql = "insert into shangqi_2 set zaotai = '%s',zaoyan = '%s',creat_time = '%s',yaji_time = '%s'" % (
                    1, number, dianhuo_time, dianhuo_time)
                try:
                    mysql_tengxun.mysql_db(sql)
                except:
                    time.sleep(0.5)
                    with open('diancifa_error.txt', 'a') as f:
                        f.write(datetime.datetime.now().strftime(
                            '%Y-%m-%d %H:%M:%S') + '_insert_data_error_' + '\n')

        excel.save(excel_filename)
        del li_dcf[0]
    win.after(500, lambda: contorl_dcf(excel_filename, li_diancifa))
def func(number):
    # 弹出对话框
    if li_diancifa[number]!=0:
        result = tkinter.messagebox.askokcancel(title='灶眼'+str(number), message='确定要点火吗？')
        if result==True:
            global li_dcf
            li_dcf.append(number)

    else:
        result = tkinter.messagebox.askokcancel(title='灶眼'+str(number), message='该灶眼还不能控制')

def get_zaoyan():

    zaotai=1
    global data_zaoyan_cunchu
    now_time=datetime.datetime.now()+datetime.timedelta(hours=-2)
    sql = "SELECT * FROM fangguo WHERE fangguo.creat_time > '%s'" % (now_time)

    data_item=[]
    while 1:
        try:
            data_item = mysql_tengxun.mysql_getdata(sql)
            break

        except:
            time.sleep(1)
            print('get fangguo data error')
            break

    data_zaoyan=[]

    if data_item:
        for i in range(len(data_item)):
            if data_item[i][2]==zaotai and data_item[i][5]==None:
                data_zaoyan.append(data_item[i][3])

    if data_zaoyan!=data_zaoyan_cunchu or data_zaoyan==[]:
        data_zaoyan_cunchu=data_zaoyan
        print('i get one data')
        li=[]
        global li_fix
        for i in range(6):
            if i not in data_zaoyan:
                li.append('Gray')
            else:
                li.append('Green')
        print(li,li_fix)
        if li_fix!=li:
            li_fix=li
            # 创建按钮
            button1 = tkinter.Button(win, text="按钮", command= lambda: func(1), width=18, height=8,bg=li[0])
            button1.place(relx=0.05, rely=0.1)

            button3 = tkinter.Button(win, text="按钮", command= lambda: func(3),width=18, height=8,bg=li[1])
            button3.place(relx=0.2, rely=0.1)

            button5 = tkinter.Button(win, text="按钮", command= lambda: func(5),width=18, height=8,bg=li[2])
            button5.place(relx=0.35, rely=0.1)

            button0 = tkinter.Button(win, text="按钮", command= lambda: func(0), width=18, height=8,bg=li[3])

            button0.place(relx=0.05, rely=0.5)

            button2 = tkinter.Button(win, text="按钮", command= lambda: func(2),width=18, height=8,bg=li[4])
            button2.place(relx=0.2, rely=0.5)

            button4 = tkinter.Button(win, text="按钮", command= lambda: func(4),width=18, height=8,bg=li[5])
            button4.place(relx=0.35, rely=0.5)

            # 创建按钮
            button7 = tkinter.Button(win, text="按钮", command= lambda: func(7), width=18,height=8,bg='red')

            button7.place(relx=0.53, rely=0.1)

            button8 = tkinter.Button(win, text="按钮", command= lambda: func(9), width=18,height=8,bg='Gray')
            button8.place(relx=0.68, rely=0.1)

            button9 = tkinter.Button(win, text="按钮", command= lambda: func(11), width=18,height=8,bg='Gray')
            button9.place(relx=0.83, rely=0.1)

            button10 = tkinter.Button(win, text="按钮", command= lambda: func(6), width=18,height=8,bg='red')

            button10.place(relx=0.53, rely=0.5)

            button11 = tkinter.Button(win, text="按钮", command= lambda: func(8), width=18,height=8,bg='Gray')
            button11.place(relx=0.68, rely=0.5)

            button12 = tkinter.Button(win, text="按钮", command= lambda: func(10), width=18,height=8,bg='Gray')
            button12.place(relx=0.83, rely=0.5)
            win.after(1000, get_zaoyan)
        else:
            win.after(1000, get_zaoyan)
    else:
        win.after(1000, get_zaoyan)

excel_filename='control_dcf.xlsx'
creat_excel(excel_filename)
win = tkinter.Tk()
win.title("左边的六眼灶")
win.geometry("1000x400+5+5")
win.after(1000,get_zaoyan)
win.after(500,lambda : contorl_dcf(excel_filename,li_diancifa))
win.mainloop()


