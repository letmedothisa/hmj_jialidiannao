# -*- coding:utf-8 -*-
import datetime
import time
import mysql_tengxun
import get_hsv
import os
import openpyxl
from openpyxl import load_workbook
import yaji_lunkuo
import del_mysql


def deal_excel(li):
    excel = load_workbook(excel_filename)
    table = excel['锅']  # 通过表名获取
    for i in range(0, 6):
        if not table['C' + str(i + 2)].value and li[i] > 5000:
            fangguo_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            fangguo_time = datetime.datetime.strptime(fangguo_time, '%Y-%m-%d %H:%M:%S')
            table['C' + str(i + 2)] = fangguo_time
            sql = "insert into fangguo set creat_time='%s',zaotai_num='%s',zaoyan='%s',guo_time='%s'" % (
            fangguo_time, int(table['A' + str(i + 2)].value), int(table['B' + str(i + 2)].value), fangguo_time)
            while 1:
                try:
                    mysql_tengxun.mysql_db(sql)
                    print('insert_one_data')
                    break
                except:
                    time.sleep(1)
                    print('insert___date__error')
                    break


        elif table['C' + str(i + 2)].value and li[i] < 5000:
            guo_long = round((datetime.datetime.now() - table['C' + str(i + 2)].value).seconds / 60, 2)

            sql = "update fangguo set guo_remove= '%s',guo_long = '%s' where guo_time = '%s' and zaoyan='%s'" % (
            datetime.datetime.now(), guo_long, table['C' + str(i + 2)].value, int(table['B' + str(i + 2)].value))

            while 1:
                try:
                    mysql_tengxun.mysql_db(sql)
                    break
                except:
                    time.sleep(1)
                    print('update__error')
                    break
            table['C' + str(i + 2)] = ''
    excel.save(excel_filename)


def get_guo(zaotai_num):
    data_item = []

    get_time = datetime.datetime.now() + datetime.timedelta(minutes=-5)

    sql = "SELECT * FROM zaotai_sign WHERE zaotai_sign.zaotai_num= '%s' and zaotai_sign.creat_time > '%s' order by zaotai_sign.creat_time DESC" % (
    zaotai_num, get_time)
    while 1:
        try:
            data_item = mysql_tengxun.mysql_getdata(sql)
            break

        except:
            time.sleep(1)
            print('get jiroudone data error')
            break

    if data_item:

        return [data_item[0][3], data_item[0][4]]
    else:
        if zaotai_num == 1:
            data_item = [1032, 30]
        else:
            data_item = [288, 65]

        return data_item


excel_filename = 'fangguo.xlsx'

if os.path.exists(excel_filename):
    os.remove(excel_filename)

# 生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
wb = openpyxl.Workbook()
# 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
ws = wb.active
# 更改工作表ws的title
ws.title = '锅'

# 对ws的单个单元格传入数据
ws['A1'] = '灶台'
ws['B1'] = '灶眼'
ws['C1'] = '拿锅的时间'
ws['A2'] = '1'
ws['A3'] = '1'
ws['A4'] = '1'
ws['A5'] = '1'
ws['A6'] = '1'
ws['A7'] = '1'
ws['A8'] = '0'
ws['A9'] = '0'
ws['A10'] = '0'
ws['A11'] = '0'
ws['A12'] = '0'
ws['A13'] = '0'
ws['B2'] = '0'
ws['B3'] = '1'
ws['B4'] = '2'
ws['B5'] = '3'
ws['B6'] = '4'
ws['B7'] = '5'
ws['B8'] = '0'
ws['B9'] = '1'
ws['B10'] = '2'
ws['B11'] = '3'
ws['B12'] = '4'
ws['B13'] = '5'
wb.save(excel_filename)

delsql_time=datetime.datetime.now()+datetime.timedelta(days=-3)

while 1:

    # 这里都是针对灶台1进行操作
    jizhun = get_guo(1)

    juli = [[-713, 413], [-658, 218], [-457, 426], [-435, 226], [-199, 433], [-208, 232]]

    gaoya_juli = [[-802, 347], [-732, 108], [-508, 347], [-472, 121], [-187, 338], [-186, 124]]

    weizhi = []

    file_path = 'D:/savepic/yaji_' + datetime.datetime.now().strftime('%Y-%m-%d') + '_pic/'

    files = os.listdir(file_path)

    next_do=False

    if len(files)>3:

        next_do=True
        file_pic = file_path + files[-2]

    if next_do==True:
        # ---------这里主要是看有没有放锅------------
        li_fangguo = []

        for i in range(len(juli)):

            x1 = juli[i][0] + jizhun[0]
            y1 = juli[i][1] + jizhun[1]
            x_long = 180
            y_long = 110

            h, s, v, place = get_hsv.gethsv(file_pic, x1, y1, x_long, y_long)
            x = []
            y = []
            k_yinse = 0
            # # 再看看是不是有火苗
            for j in range(len(h)):

                if (260 > v[j] > 145 and 75>s[j]>0 and 120>h[j]>10):
                    x.append(place[j][0])
                    y.append(place[j][1])
                    k_yinse += 1

            li_fangguo.append(k_yinse)

        print(li_fangguo)

        deal_excel(li_fangguo)


        # 这里看看灶眼的位置
        gaoya = []
        for i in range(len(li_fangguo)):
            if li_fangguo[i]>5000:
                li_zaoyan = []
                li_zaoyan.append(gaoya_juli[i][0] + jizhun[0])
                li_zaoyan.append(gaoya_juli[i][1] + jizhun[1])
                li_zaoyan.append(280)
                li_zaoyan.append(180)

                try:
                    outline = yaji_lunkuo.main(li_zaoyan, file_pic, 130, 130, 230)
                    for j in range(len(outline)):
                        x_long = outline[j][2] - outline[j][1]
                        y_long = outline[j][4] - outline[j][3]
                        center_x=int((outline[j][1]+outline[j][2])/2)
                        center_y=int((outline[j][3]+outline[j][4])/2)
                        total_long = outline[j][0]

                        yuanzhoulv = (total_long / ((x_long + y_long) * 2))
                        changkuanbi = y_long / x_long
                        if 0.66 < yuanzhoulv < 1 and 1.1 > changkuanbi > 0.65:
                            sql = "insert into gaoya_position set creat_time='%s',zaotai_num='%s',zaoyan='%s',min_x='%s',max_x='%s',min_y='%s',max_y='%s',center_x='%s',center_y='%s'" % (
                                datetime.datetime.now(),1,i,outline[j][1],outline[j][2],outline[j][3],outline[j][4],center_x,center_y)
                            while 1:
                                try:
                                    mysql_tengxun.mysql_db(sql)
                                    print('insert_one_data')
                                    break
                                except:
                                    time.sleep(1)
                                    print('insert___date__error')
                                    break
                        else:
                            file_error='./file_error/'+datetime.datetime.now().strftime('%H-%M-%S')+str(i)+'.png'
                            os.rename(file_pic,file_error)
                except:
                    time.sleep(1)
                    print('get_lunkuo_error')

    #删除sql数据
    if (datetime.datetime.now() - delsql_time).days > 3:

        now_time = datetime.datetime.now() + datetime.timedelta(days=-3)

        sql = "DELETE FROM fangguo WHERE creat_time <'%s'" % now_time

        del_mysql.del_mysql(sql)

        sql = "DELETE FROM gaoya_position WHERE creat_time <'%s'" % now_time

        del_mysql.del_mysql(sql)

        delsql_time = datetime.datetime.now()

    time.sleep(15)