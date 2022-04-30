# -*- coding:utf-8 -*-
import os
import time
import datetime
import sys
import openpyxl
from openpyxl import load_workbook
import mysql_tengxun
import jietu_getrgb
from tkinter import messagebox
import getall_basedata
import del_file


def get_xy(zaotai, zaoyan, get_time, li_standard):
    # 计算出x,y,xlong,ylong
    biaoji = get_guo(zaotai)

    data_item = []

    sql = "SELECT * FROM gaoya_position WHERE gaoya_position.zaotai_num='%s' and gaoya_position.zaoyan='%s' and gaoya_position.creat_time>'%s'" % (
        zaotai, zaoyan, get_time)
    while 1:
        try:
            data_item = mysql_tengxun.mysql_getdata(sql)
            break

        except:
            time.sleep(1)
            print('get gaoya_position data error')
            break

    if data_item:

        x = data_item[-1][4]
        y = data_item[-1][6]
        x_long = data_item[-1][5] - data_item[-1][4]
        y_long = data_item[-1][7] - data_item[-1][6]

    else:

        x = int(li_standard[zaoyan][0]) + int(biaoji[0]) - 50
        y = int(li_standard[zaoyan][1]) + int(biaoji[1]) - 50
        x_long = 100
        y_long = 100

    li_xy = []
    li_xy.append(x)
    li_xy.append(y)
    li_xy.append(x_long)
    li_xy.append(y_long)

    return x, y, x_long, y_long, li_xy


def creat_excel(excel_filename):
    if os.path.exists(excel_filename):
        os.remove(excel_filename)
        time.sleep(0.5)

    # 生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
    wb = openpyxl.Workbook()
    # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
    ws = wb.active
    # 更改工作表ws的title
    ws.title = '压鸡上汽时间'
    ws['A1'] = '灶眼'
    ws['B1'] = '点火时间'
    ws['C1'] = '最新判断时间'
    ws['D1'] = '初始值'
    ws['E1'] = '上汽值'
    ws['F1'] = '上汽时间'
    ws['G1'] = '上汽次数'
    ws['H1'] = '是否已经截图'
    ws['I1'] = '是否上传初始值'
    ws['J1'] = '是否上传上汽值'
    ws['A2'] = '0'
    ws['A3'] = '1'
    ws['A4'] = '2'
    ws['A5'] = '3'
    ws['A6'] = '4'
    ws['A7'] = '5'
    ws['G2'] = '0'
    ws['G3'] = '0'
    ws['G4'] = '0'
    ws['G5'] = '0'
    ws['G6'] = '0'
    ws['G7'] = '0'
    wb.save(excel_filename)


def del_pic(file_path, cutpic_path, del_pictime, del_cutpictime):
    # 按照排序删除多余的图片
    if (datetime.datetime.now() - del_pictime).seconds > pic_savetime * 60:
        # 把清除图片的时间赋值回来
        dir_list = os.listdir(file_path)
        # 如果有文件，就挨个获取文件的最后修改时间，如果
        if dir_list:
            if len(dir_list) > 1500:
                for i in range(len(dir_list) - 1000):
                    file_pic = file_path + dir_list[i]
                    os.remove(file_pic)
                del_pictime = datetime.datetime.now()

    if (datetime.datetime.now() - del_cutpictime).days >= cutpic_savetime:
        del_cutpictime = datetime.datetime.now()
        del_file.del_file(cutpic_path, 2, 0, 0)

    return del_pictime, del_cutpictime


def get_guo(zaotai_num):
    data_item = []

    get_time = datetime.datetime.now() + datetime.timedelta(minutes=-5)

    sql = "SELECT * FROM zaotai_sign WHERE zaotai_sign.zaotai_num= '%s' and zaotai_sign.creat_time > '%s' order by zaotai_sign.creat_time DESC" % (
        zaotai_num, get_time)
    while 1:
        try:
            data_item = mysql_tengxun.mysql_getdata(sql)
            break

        except:
            time.sleep(1)
            print('get zaotai_sign data error')
            break

    if data_item:

        return [data_item[0][3], data_item[0][4]]
    else:
        if zaotai_num == 1:
            data_item = [1032, 30]
        else:
            data_item = [288, 65]

        return data_item


def get_sum(file_path, li_outline):
    get_paixu = False

    dir_list = os.listdir(file_path)

    try:
        dir_list = sorted(dir_list, key=lambda x: os.path.getmtime(os.path.join(file_path, x)),
                          reverse=True)
        get_paixu = True

    except:
        print('get_paixu_error')

    sum = 0

    if get_paixu == True:

        if len(dir_list) > 19:

            for i in range(1, 19):

                sum0 = 0

                file_pic = file_path + dir_list[i]

                x = li_outline[0]
                y = li_outline[1]
                x_long = li_outline[2]
                y_long = li_outline[3]

                try:
                    h, s, v, place = jietu_getrgb.getrgb(file_pic, x, y, x_long, y_long)
                except:
                    print('get_hsv_eroor')
                    sum = 0
                    break

                if i > 1:

                    for o in range(0, x_long * y_long):
                        sum0 += abs(int(h[o]) - int(h_before[o]))
                        sum0 += abs(int(s[o]) - int(s_before[o]))
                        sum0 += abs(int(v[o]) - int(v_before[o]))
                    sum0 = round(sum0 / (x_long * y_long), 2)
                    sum += sum0

                h_before = h
                s_before = s
                v_before = v
    return sum


def get_data(li2, li3, excel_filename, zaotai):
    data = []
    # 去拿第一个六眼灶的数据
    while 1:
        try:
            data = getall_basedata.get_data(0.5, 'shangqi_2', 'yaji_time')
            break
        except:
            print('error')
            time.sleep(1)
            continue

    if data:
        for i in range(len(data)):

            # 如果找到点火数据时候的处理
            if int(data[i][2]) == zaotai and data[i][5] == None and data[i][0] not in li2:
                li2.append(data[i][0])

                # 把符合条件的灶眼拿出来
                zaoyan = int(data[i][3])
                # 获取点火的实际时间，并写入excel
                yaji_time = data[i][4]

                # 打开excel文件，并找到灶眼
                excel = load_workbook(excel_filename)
                # 获取sheet：
                table = excel['压鸡上汽时间']  # 通过表名获取

                table['B' + str(zaoyan + 2)] = yaji_time
                table['C' + str(zaoyan + 2)] = yaji_time

                excel.save(excel_filename)

            # 找到关火数据
            if int(data[i][2]) == zaotai and data[i][5] != None and data[i][0] not in li3:

                li3.append(data[i][0])
                # 把符合条件的灶眼拿出来

                zaoyan = int(data[i][3])
                # 获取点火的实际时间，并写入文本

                excel = load_workbook(excel_filename)
                # 获取sheet：
                table = excel['压鸡上汽时间']  # 通过表名获取

                # 有一次完整的点火，并且上汽了
                if table['B' + str(zaoyan + 2)].value and table['F' + str(zaoyan + 2)].value and int(
                        table['G' + str(zaoyan + 2)].value) >= 3:

                    shangqi_long = round(
                        (table['F' + str(zaoyan + 2)].value - table['B' + str(zaoyan + 2)].value).seconds / 60, 2)
                    sql = "update shangqi_2 set shangqi_data = '%s',chushi_data = '%s',shangqi_time = '%s',shangqi_num = '%s',shangqi_long = '%s' where yaji_time='%s'" % (
                        table['E' + str(zaoyan + 2)].value, table['D' + str(zaoyan + 2)].value,
                        table['F' + str(zaoyan + 2)].value,
                        table['G' + str(zaoyan + 2)].value, shangqi_long, table['B' + str(zaoyan + 2)].value)

                    while 1:
                        try:
                            mysql_tengxun.mysql_db(sql)
                            break
                        except:
                            print("update error")
                            time.sleep(1)
                            continue

                for k in range(ord('B'), ord('K')):
                    table[chr(k) + str(zaoyan + 2)] = ''
                table['G' + str(zaoyan + 2)] = 0

                excel.save(excel_filename)

    return li2, li3


# --------------------------------------------------------进行所有的初始化配置--------------------------------
# 所有初始化的项目
li2 = []
li3 = []
pic_savetime = 0
judge_inteval = 5
del_pictime = datetime.datetime.now()
del_cutpictime = datetime.datetime.now()

# 获取exe信息
a = sys.executable
b = a.split('\\')
get_dowhat = False
for i in range(len(b)):
    if 'exe' in b[i]:
        dowhat = b[i].split('.')[0]
        get_dowhat = True

dowhat = 'yaji_getresult'
get_dowhat = True

if get_dowhat == True:
    dowhat_prehalf = dowhat.split('_')[0]
    with open('config.txt', 'r') as f:
        a = f.readlines()
        for i in range(len(a)):
            b = a[i].split(':')
            if b[0] == dowhat:
                zaoyan_standard = b[1].split('=')[1].replace(' ', '')
                zaotai = int(b[2].replace('\n', '').split('=')[1])
                sum_change = int(b[3].replace('\n', '').split('=')[1])

    with open('config.txt', 'r') as f:
        a = f.readlines()
        for i in range(len(a)):
            b = a[i].split(' ')
            if b[0] == dowhat_prehalf:
                pic_savetime = int(b[4].replace("\n", "").replace("\r", "").split('=')[1])
                picture_path = b[3].replace("\n", "").replace("\r", "").split('=')[1]

            if b[0] == 'cutpic_path':
                cutpic_path = b[1].replace("\n", "").replace("\r", "").replace("]", "")
                cutpic_savetime = int(b[2].replace("\n", "").split('=')[1])

        if not os.path.exists(cutpic_path):
            os.makedirs(cutpic_path)

    zaoyan_standard = zaoyan_standard.split(',')

    li_standard = []
    for i in range(int(len(zaoyan_standard) / 2)):
        li_zaoyan_1 = []
        li_zaoyan_1.append(int(zaoyan_standard[i * 2].replace('[', '').replace(']', '')))
        li_zaoyan_1.append(int(zaoyan_standard[i * 2 + 1].replace('[', '').replace(']', '')))
        li_standard.append(li_zaoyan_1)


else:
    messagebox.showinfo('提示', '没有找到dowhat信息，请注意！')

# ----------------------------------------建新的excel文件-------------------------

excel_filename = dowhat + '.xlsx'
creat_excel(excel_filename)

# -------------------------------------------------------开始干活，先把多余的图片文件删除掉----------------------------------------
while 1:

    file_path = picture_path + datetime.datetime.now().strftime('%Y-%m-%d') + '_pic/'
    if not os.path.exists(file_path):
        os.makedirs(file_path)

    del_pictime, del_cutpictime = del_pic(file_path, cutpic_path, del_pictime, del_cutpictime)

    li2, li3 = get_data(li2, li3, excel_filename, zaotai)

    files = os.listdir(file_path)

    get_filepic = False
    if len(files) > 2:
        file_pic = file_path + files[-2]
        get_filepic = True

    # ------------------------------------------------------先确定x,y,xlong,ylong--------------------------------
    for i in range(2, 8):

        zaoyan = i - 2
        # 打开excel表格
        excel = load_workbook(excel_filename)
        # 获取sheet：
        table = excel['压鸡上汽时间']  # 通过表名获取

        if table['B' + str(i)].value:

            if not table['C' + str(i)].value:
                table['C' + str(i)] = datetime.datetime.now()

            if table['E' + str(i)].value:
                judge_inteval=10

            # 去获取x,y,xlong,ylong
            x, y, x_long, y_long, li_xy = get_xy(zaotai, zaoyan, table['B' + str(i)].value, li_standard)

            # -------------------------------------计算初始值---------------------------------：

            if (datetime.datetime.now() - table['C' + str(i)].value).seconds > 20 and 60 < (
                    datetime.datetime.now() - table['B' + str(i)].value).seconds < 360:

                # 把高压帽的位置进行截图
                if not table['H' + str(i)].value and get_filepic == True:
                    table['H' + str(i)] = 1
                    file_save_pic = cutpic_path + str(zaotai) + '_' + str(
                        i - 2) + '_' + datetime.datetime.now().strftime(
                        '%Y-%m-%d') + '_' + datetime.datetime.now().strftime('%H-%M-%S') + '_ceshi.png'
                    jietu_getrgb.cutpic(file_pic, x, y, x_long, y_long, file_save_pic)

                # 去获取sum值
                sum = get_sum(file_path, li_xy)

                sum=int(sum)

                if not table['I' + str(i)].value:

                    sql = "update shangqi_2 set chushi_list='%s' where shangqi_2.yaji_time='%s'" % (
                        sum, table['B' + str(i)].value)

                    table['I' + str(i)] = 15
                else:
                    list_sum = '-' + str(sum)

                    sql = "update shangqi_2 set chushi_list=CONCAT(chushi_list,%s) where shangqi_2.yaji_time='%s'" % (
                        list_sum, table['B' + str(i)].value)
                while 1:
                    try:
                        mysql_tengxun.mysql_db(sql)
                        break
                    except:
                        print('updata_chushidata_list_error')
                        break

                while 1:
                    try:
                        mysql_tengxun.mysql_db(sql)
                        break
                    except:
                        print('updata_chushidata_list_error')
                        break

                if sum > 0:

                    print('-------------------------本次初始值为，%d,,灶眼是%d' % (sum, zaoyan))
                    if not table['D' + str(i)].value:
                        table['D' + str(i)] = sum
                    else:
                        if int(table['D' + str(i)].value) > sum:
                            table['D' + str(i)] = sum
                            print('我更新了一次初始值--------------为，%d,,更新以后的值是%d' % (
                                sum, int(table['D' + str(i)].value)))
                    table['C' + str(i)] = datetime.datetime.now()


            # ------------------------------------- 判断在9分钟以后，开始判断是否已经上汽--------------------------------------------------------------------

            elif (datetime.datetime.now() - table['C' + str(i)].value).seconds > judge_inteval and (
                    datetime.datetime.now() - table['B' + str(i)].value).seconds > 500:

                # 判断如果初始值为空，则不走后续流程，在初始值处，填写0
                if not table['D' + str(i)].value or table['D' + str(i)].value == 0:
                    table['D' + str(i)] = 0

                else:
                    sum = get_sum(file_path, li_xy)

                    sum=int(sum)

                    if not table['J' + str(i)].value:

                        sql = "update shangqi_2 set shangqi_list='%s' where shangqi_2.yaji_time='%s'" % (
                        sum, table['B' + str(i)].value)

                        table['J' + str(i)] = 1
                    else:
                        list_sum = '-' + str(sum)

                        sql = "update shangqi_2 set shangqi_list=CONCAT(shangqi_list,%s) where shangqi_2.yaji_time='%s'" % (
                            list_sum, table['B' + str(i)].value)
                    while 1:
                        try:
                            mysql_tengxun.mysql_db(sql)
                            break
                        except:
                            print('updata_chushidata_list_error')
                            break

                    if sum > 0:
                        print('-----------本次上汽值为，%d,,灶眼是%d' % (sum, zaoyan))
                        if (sum - table['D' + str(i)].value > sum_change) and sum < 850:

                            if table['E' + str(i)].value:
                                if sum > table['E' + str(i)].value:
                                    table['E' + str(i)] = sum
                                    print('我更新了一次上汽值--------------为，%d,,更新以后的值是%d' % (
                                        sum, int(table['E' + str(i)].value)))
                            else:
                                table['E' + str(i)] = sum

                            table['G' + str(i)] = int(table['G' + str(i)].value) + 1

                            if table['G' + str(i)].value == 5:
                                table['F' + str(i)] = datetime.datetime.now()

                                # 把已经上汽的情况，传到服务器
                                sql = "update shangqi_2 set shangqi_time = '%s' where yaji_time = '%s'" % (
                                    str(datetime.datetime.now()), table['B' + str(i)].value)
                                while 1:
                                    try:
                                        mysql_tengxun.mysql_db(sql)
                                        break
                                    except:

                                        print("update error")
                                        time.sleep(3)
                                        break
                            elif table['G' + str(i)].value > 5:
                                shangqi_delay = 15
                        elif sum - table['D' + str(i)].value <= sum_change and int(
                                table['G' + str(i)].value) > 0:
                            table['G' + str(i)] = int(table['G' + str(i)].value) - 1

                table['C' + str(i)] = datetime.datetime.now()
        excel.save(excel_filename)

    if len(li2) > 50:
        del li2[0]
    if len(li3) > 50:
        del li3[0]
    time.sleep(2)
    print(datetime.datetime.now())